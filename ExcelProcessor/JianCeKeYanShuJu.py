import xlrd
import pandas as pd
import openpyxl
import os


def read_excel(file_path):
    '''
    从excel中读取数据
    :param file_path: 文件路径
    :return: 以每天日期为key，测量值为value的字典。注意因为字典的key不能相同，返回值可能合并了key相同的项，可以考虑返回list
    '''
    workbook = xlrd.open_workbook(file_path)
    #sheet = workbook.sheet_by_name('静力水准仪')
    sheet = workbook.sheet_by_name('地连墙水平位移')
    index = []
    data = []
    #for i in range(1, sheet.nrows):
    for i in range(2, 52):
        if sheet.row(i)[0].ctype == 3:
            mode = workbook.datemode
            t = xlrd.xldate_as_datetime(sheet.row_values(i)[0], mode)
            index.append(str(t))
        else:
            index.append(sheet.row_values(i)[0])
        data.append(sheet.row_values(i)[1:30])
    if len(set(index)) != len(index):
        print('*-'*60, '\n注意：已经合并了某些项\n', '-*'*60)
    return dict(zip(index, data))


def get_data_mean(df_data):
    '''
    计算同一天内的平均值
    :param df_data: DataFrame数据，每天数据为一行，一天有多行
    :return: 返回含每天平均值的dataframe，一天只有一行，行数据为平均值
    '''
    index_list = list(set(df_data.index))
    index_list.sort()
    result_list = []
    for index in index_list:
        mean_data = df_data.ix[index]
        if isinstance(mean_data, pd.DataFrame):
            result_list.append(mean_data.mean(axis = 0))
        else:
            result_list.append(mean_data)
    return pd.DataFrame(result_list, index=index_list)


def data_to_excel(df_data, excel_path, sheet_name):
    writer = pd.ExcelWriter(excel_path)
    df_data.to_excel(writer, sheet_name)
    writer.close()


def add_data_to_excel(df_data, excel_path, sheet_name):
    '''
    往excel中增加数据
    :param df_data: 要写的数据
    :param excel_path: 路径
    :param sheet_name: 表名
    :return: 无
    '''
    if os.path.isfile(excel_path):
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.create_sheet(sheet_name)
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
    data_list = []
    l = list(df_data.columns)
    l.insert(0, '深度')
    data_list.append(l)
    for index, row in df_data.iterrows():
        t = list(row)
        t.insert(0, index)
        data_list.append(t)

    for d in data_list:
        worksheet.append(d)
    workbook.save(excel_path)


if  __name__ == '__main__':
    data = read_excel(r'C:\Users\Administrator\Desktop\test\数据汇总(1月6日）.xls')
    # print(data)
    # df = pd.DataFrame(data, index=['JL3', 'JL2', 'JL1']).T
    #df = pd.DataFrame(data, index=['0m', '3m', '6m', '9m', '12m', '15m', '18m', '21m', '24m', '27m', '30m']).T
    df = pd.DataFrame(data, index=['0m', '1m', '2m', '3m', '4m', '5m', '6m', '7m', '8m', '9m', '10m', '11m', '12m', '13m', '14m', '15m', '16m', '17m', '18m', '19m', '20m', '21m', '22m', '23m', '24m', '25m', '26m', '27m', '28m']).T
    my_df = df.rename(index = lambda s : s[:10])
    #print(my_df)
    result = get_data_mean(my_df)
    #data_to_excel(result, r'C:\Users\Administrator\Desktop\test\result.xlsx', '静力水准仪')
    add_data_to_excel(result, r'C:\Users\Administrator\Desktop\test\reslult.xlsx', 'CX3')