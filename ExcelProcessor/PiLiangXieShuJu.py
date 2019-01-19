import openpyxl
import glob
import os
import pandas as pd


def write_data(excel_path, sheet_name, data_list, start_row, start_col):
    '''
    向已有excel中写入二维列表数据
    :param excel_path:
    :param sheet_name:
    :param data_list: 二维列表
    :param start_row: excel中从1开始的行数
    :param start_col: excel中从1开始的列数
    :return:
    '''
    workbook = openpyxl.load_workbook(excel_path)
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        print('找到表：' + sheet_name)
    else:
        worksheet = workbook.create_sheet(sheet_name)
        print('未找到表：' + sheet_name + '， 已自动创建')

    print('写入数据：' + str(data_list))
    for i in range(len(data_list)):
        for j in range(len(data_list[i])):
            cell = worksheet.cell(start_row + i, start_col + j, data_list[i][j])

    workbook.save(excel_path)
    print('数据写入完毕!\n')



def get_all_files(dir_path, ext = '*.*'):
    '''
    得到指定目录下的所有文件列表
    :param dir_path: 文件夹路径，如C:\\Users\\Administrator\\Desktop\\test
    :param ext: 过滤文件名, 如'*.txt'
    :return: 文件完整路径列表
    '''
    if dir_path.endswith('\\'):
        filter_ext = dir_path +ext
    else:
        filter_ext = dir_path + '\\' +ext
    files = []
    for filename in glob.glob(filter_ext):
        files.append(filename)
    return files


def modify_CX():
    '''
    将所有CX第二列数据改为0
    :return:
    '''
    files = get_all_files(r'C:\Users\Administrator\Desktop\新建文件夹', '*.xlsx')
    for f in files:
        print('正在处理：' + f)
        workbook = openpyxl.load_workbook(f)
        for sheet in workbook.sheetnames:
            if  'CX' in sheet:
                print('正在处理sheet：' + sheet)
                worksheet = workbook[sheet]
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row, 2, 0)
        workbook.save(f)


def modify_sheet():
    '''
    修改sheet名，删除sheet
    :return:
    '''
    files = get_all_files(r'C:\Users\Administrator\Desktop\新建文件夹', '*.xlsx')
    for f in files:
        print('正在处理：' + f)
        workbook = openpyxl.load_workbook(f)
        qingxie_sheet = workbook['科研监测项目_地墙倾斜']
        workbook.remove(qingxie_sheet)
        sheet = workbook['科研监测项目_光栅地墙接缝']
        sheet.title = '科研光栅地墙接缝项目_光栅地墙接缝'
        sheet = workbook['科研监测项目_静力水准']
        sheet.title = '科研静力水准项目_静力水准'
        sheet = workbook['科研监测项目_土压力']
        sheet.title = '科研土压力项目_土压力'
        workbook.create_sheet('科研立柱桩内力项目_立柱桩内力')
        workbook.create_sheet('科研土体位移项目_土体位移')
        workbook.create_sheet('科研支撑轴力项目_支撑轴力')
        workbook.save(f)


def XieShuJu1():
    files = get_all_files(r'C:\Users\Administrator\Desktop\新建文件夹', '*.xlsx')
    for f in files:
        print('正在处理：' + f)
        write_data(f, '科研土压力项目_土压力', [['点号', '初始测值(MPa)', '本次测值(MPa)'], ['KY-TYL1'], ['KY-TYL2']], 1, 1)
        write_data(f, '科研立柱桩内力项目_立柱桩内力', [['点号', '初始测值(kN)', '本次测值(kN)'], ['KY-LZ']], 1, 1)
        write_data(f, '科研土体位移项目_土体位移', [['点号', '初始测值(mm)', '本次测值(mm)'], ['KY-TTWY']], 1, 1)
        write_data(f, '科研支撑轴力项目_支撑轴力', [['点号', '初始测值(kN)', '本次测值(kN)'], ['KY-ZL'], ['KY-ZL-2'], ['KY-ZL-3'], ['KY-ZL-4'], ['KY-ZL-5']], 1, 1)
        write_data(f, '科研静力水准项目_静力水准', [['点号', '初始测值(mm)', '本次测值(mm)']], 1, 1)
        write_data(f, 'KY-CX1', [['深度(m)', '初始测值(mm)', '本次测值(mm)']], 1, 1)
        write_data(f, 'KY-CX2', [['深度(m)', '初始测值(mm)', '本次测值(mm)']], 1, 1)
        write_data(f, 'KY-CX3', [['深度(m)', '初始测值(mm)', '本次测值(mm)']], 1, 1)
        write_data(f, 'KY-CX4', [['深度(m)', '初始测值(mm)', '本次测值(mm)']], 1, 1)


def CopyData():
    '''
    将某列数据拷贝到另一列
    :return:
    '''
    def copy(workbook, sheet_names, start_row):
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            for r in range(start_row, worksheet.max_row + 1):
                worksheet.cell(r, 2, worksheet.cell(r, 3).value)

    files = get_all_files(r'C:\Users\Administrator\Desktop\新建文件夹', '*.xlsx')
    for f in files:
        print('正在处理：' + f)
        workbook = openpyxl.load_workbook(f)
        copy(workbook, ['科研监测项目_地墙倾斜', '科研监测项目_静力水准', '科研监测项目_土压力', 'KY-CX3', 'KY-CX4'], 2)
        copy(workbook, ['支撑轴力_第一道砼支撑', '支撑轴力_第一道钢支撑', '支撑轴力_第二道钢支撑', '支撑轴力_第三道钢支撑', '支撑轴力_第四道钢支撑'], 11)
        workbook.save(f)


def modify_zhichengzhouli():
    files = get_all_files(r'C:\Users\Administrator\Desktop\新建文件夹', '*.xlsx')
    for f in files:
        print('正在处理：' + f)
        workbook = openpyxl.load_workbook(f)
        keyan_sheet = workbook['科研支撑轴力项目_支撑轴力']
        diyidao = workbook['支撑轴力_第一道砼支撑']
        keyan_sheet.cell(2, 2, diyidao['B11'].value)
        keyan_sheet.cell(2, 3, diyidao['B11'].value)
        diyidao = workbook['支撑轴力_第一道钢支撑']
        keyan_sheet.cell(3, 2, diyidao['B11'].value)
        keyan_sheet.cell(3, 3, diyidao['B11'].value)
        diyidao = workbook['支撑轴力_第二道钢支撑']
        keyan_sheet.cell(4, 2, diyidao['B11'].value)
        keyan_sheet.cell(4, 3, diyidao['B11'].value)
        diyidao = workbook['支撑轴力_第三道钢支撑']
        keyan_sheet.cell(5, 2, diyidao['B11'].value)
        keyan_sheet.cell(5, 3, diyidao['B11'].value)
        diyidao = workbook['支撑轴力_第四道钢支撑']
        keyan_sheet.cell(6, 2, diyidao['B11'].value)
        keyan_sheet.cell(6, 3, diyidao['B11'].value)
        workbook.save(f)


def get_tutiweiyi():
    data = pd.read_excel(r'C:\Users\Administrator\Desktop\数据整理result.xlsx', '钢支撑轴力')
    for f in data.index:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'KY-TTWY'
        worksheet['A1'] = '深度(m)'
        worksheet['B1'] = '初始测值(mm)'
        worksheet['C1'] = '本次测值(mm)'
        workbook.save(os.path.join(r'C:\Users\Administrator\Desktop\土体位移', '土体位移'+f+'.xlsx'))

def fill_tutiweiyi_data():
    def getname(s):
        name = os.path.basename(s)
        shortname = os.path.splitext(name)[0]
        return shortname[-10:]

    files = get_all_files(r'C:\Users\Administrator\Desktop\土体位移', '*.xlsx')
    datename = list(map(getname, files))
    print(datename)

    deepth_data = [[str(i) + 'm'] for i in range(24)]
    deepth_data.append(['23.5m'])
    deepth_data.append(['24.5m'])
    zero_data = [[0] for i in range(26)]

    data = pd.read_excel(r'C:\Users\Administrator\Desktop\数据整理result.xlsx', '土体深部位移监测')
    for f in datename:
        if f in data.index:
            data_for_write = list(data.ix[f, 0:26])
            data_for_write = list(map(lambda s: [s], data_for_write))
            data_for_write.reverse()
            inx = datename.index(f)
            print('正在处理：' + files[inx])
            write_data(files[inx], 'KY-TTWY', data_for_write, 2, 3)
            write_data(files[inx], 'KY-TTWY', deepth_data, 2, 1)
            write_data(files[inx], 'KY-TTWY', zero_data, 2, 2)


def XieShuJu2():
    def getname(s):
        name = os.path.basename(s)
        shortname = os.path.splitext(name)[0]
        return shortname[-5:].replace('.', '-')

    files = get_all_files(r'C:\Users\Administrator\Desktop\新建文件夹', '*.xlsx')
    datename = list(map(getname, files))
    print(datename)

    data = pd.read_excel(r'C:\Users\Administrator\Desktop\数据整理result.xlsx', '钢支撑轴力').rename(index=lambda s:s[5:])
    for f in datename:
        if f in data.index:
            data_for_write = list(data.ix[f, 0:26])
            data_for_write = list(map(lambda s: [s], data_for_write))
            # data_for_write.reverse()
            inx = datename.index(f)
            print('正在处理：' + files[inx])
            write_data(files[inx], '科研土体位移项目_土体位移', data_for_write, 2, 3)


if __name__ == '__main__':
    # data = pd.read_excel(r'C:\Users\Administrator\Desktop\数据整理result.xlsx', '土体深部位移监测').rename(index=lambda s:s[5:])
    # print(data.ix['11-20',0:26])
    # XieShuJu2()
    # XieShuJu1()
    # CopyData()
    # modify_sheet()
    # modify_zhichengzhouli()
    # get_tutiweiyi()
    fill_tutiweiyi_data()

