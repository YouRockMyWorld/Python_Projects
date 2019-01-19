import openpyxl
import openpyxl.styles


def filter_row_data_1(row_data):
    '''
    有线电视、电信 行过滤器
    :param row_data:
    :return: bool
    '''
    data = list(row_data)

    #过滤掉起点终点相同的
    if data[1] == data[2]:
        return False

    #过滤掉起始点都是大数的
    if data[1] and int(data[1]) >= 1e13 or data[2] and int(data[2]) > 1e13:
        return False

    #过滤掉管线规则错误的或者小于100的
    if not data[7]:
        return False
    if '×' in str(data[7]):
        if data[7] == '×':
            return False
        length, width = data[7].split('×')
        if int(length) < 100 or int(width) < 100:
            return False
    elif data[7] < 100:
        return False

    #过滤掉埋深小于0的
    # if data[12] < 0 or data[13] < 0:
    #     return False

    return True


def filter_row_data_2(row_data):
    '''
    有线电视、电信 行过滤器
    :param row_data:
    :return: bool
    '''
    data = list(row_data)

    #过滤掉起点终点相同的
    if data[1] == data[2]:
        return (0, '物探点号、连接点号相同')

    #过滤掉起始点都是大数的
    if data[1] and int(data[1]) >= 1e13 or data[2] and int(data[2]) > 1e13:
        return (1, '物探点号或者连接点号为无意义大数字')

    #过滤掉管线规则错误的或者小于100的
    if data[7] is None:
        return (0, '规格为空值')
    if '×' in str(data[7]):
        if data[7] == '×':
            return (0, '规格错误')
        length, width = data[7].split('×')
        if int(length) < 100 or int(width) < 100:
            return (2, '规格小于100')
    elif data[7] < 100:
        return (2, '规格小于100')

    #过滤掉埋深小于0的
    # if data[12] < 0 or data[13] < 0:
    #     return False

    return (3, '数据正常')


def test1():
    workbook = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\湖.xlsx')
    name_list = workbook.sheetnames[3:]
    print(name_list)
    for name in name_list:
        print(name)
        worksheet = workbook[name]
        index = 1
        for row in worksheet.values:
            # print(index)
            if 5 <= index % 37 <=35 and filter_row_data_1(row):
                print('%d : %s | %s'  % (index, filter_row_data_1(row), row))
            index+=1
    # worksheet = workbook['供电']


def xie_yichang_yuanyin():
    workbook = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\湖.xlsx')
    name_list = workbook.sheetnames[3:]

    font = openpyxl.styles.Font(name='微软雅黑', size=8)
    align = openpyxl.styles.Alignment(vertical='center', horizontal='center')
    red_fill = openpyxl.styles.PatternFill(fgColor='FF0000', fill_type='solid')
    yellow_fill = openpyxl.styles.PatternFill(fgColor='FFFF00', fill_type='solid')
    green_fill = openpyxl.styles.PatternFill(fgColor='00B050', fill_type='solid')
    blue_fill = openpyxl.styles.PatternFill(fgColor='00B0F0', fill_type='solid')
    thin = openpyxl.styles.Side(color='000000', border_style='thin')
    border = openpyxl.styles.Border(top=thin, left=thin, bottom=thin, right=thin)

    for worksheet in workbook.worksheets[3:]:
        index = 1
        print(worksheet.title)
        for row in worksheet.values:
            if 5 <= index % 37 <= 35:
                info = filter_row_data_2(row)
                cell = worksheet.cell(index, 22, info[1])
                cell.font = font
                cell.alignment = align
                cell.border = border
                if info[0] == 0:
                    cell.fill = red_fill
                elif info[0] == 1:
                    cell.fill = yellow_fill
                elif info[0] == 2:
                    cell.fill = blue_fill
                elif info[0] == 3:
                    cell.fill = green_fill
            index+=1
        worksheet.column_dimensions['V'].width = 30

    workbook.save(r'C:\Users\Administrator\Desktop\管线数据校验结果.xlsx')




if __name__ == '__main__':
    # workbook = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\供电.xlsx')
    # worksheet = workbook['供电']
    # # for row in worksheet.iter_rows(min_row=1, max_row=50):
    #     # print(list(map(lambda v:v.value, row)))
    #
    # # j = 0
    # # for i in worksheet.values:
    # #     if j <= 10:
    # #         print(i)
    # #         j+=1
    # index = 1
    # for row in worksheet.values:
    #     # print(index)
    #     if filter_row_data_1(row):
    #         print('%d : %s | %s'  % (index, filter_row_data_1(row), row))
    #     index+=1
    # test1()
    xie_yichang_yuanyin()